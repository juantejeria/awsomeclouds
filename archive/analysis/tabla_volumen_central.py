"""
Tabla del VOLUMEN CENTRAL: franja de ±1/8 de la distancia cruz↔anca tomada
alrededor del PUNTO MEDIO entre ambos diámetros.

Para cada individuo (modelos _v8 sin cresta):
  xMid  = (xCruz + xAnca) / 2
  dist  = |xAnca - xCruz|
  franja = [xMid - dist/8,  xMid + dist/8]   (ancho total = 2/8·dist = 1/4 de dist)
  volumen = integral de la sección transversal COMPLETA (sin piso) en esa franja.

Mismo slicing que el visor 3D. Salida: CSV + Excel (una fila por individuo).
Uso: python tabla_volumen_central.py
"""
import csv
import glob
import os
import json
import numpy as np
import trimesh
from tabla_volumen_corte import slice_section, clipped_volume_liters, V8

OCULTOS = {'000_392', '000_392A', '000_448A', '000_459'}  # 12junio ocultos
FRAC = 1.0 / 8.0  # 1/8 de la distancia a cada lado


def _clamp(v, lo, hi):
    return min(max(float(v), lo), hi)


def main():
    rows = []
    for lab, d in V8.items():
        for ind in sorted(glob.glob(d + '/*')):
            if not os.path.isdir(ind):
                continue
            name = os.path.basename(ind)
            if lab == '12junio' and name in OCULTOS:
                continue
            ply = glob.glob(ind + '/*_3d.ply')
            rj = glob.glob(ind + '/*_resumen.json')
            if not ply or not rj:
                continue
            meta = json.load(open(rj[0]))
            m = trimesh.load(ply[0], process=False)
            V = np.asarray(m.vertices, float)
            F = np.asarray(m.faces, int)
            if len(F) == 0:
                continue
            x = V[:, 0]
            xmin, xmax = x.min(), x.max()
            L = xmax - xmin
            fr = (meta.get('barril_dir') != 'left')
            xf = xmax if fr else xmin
            xr = xmin if fr else xmax
            cf = meta.get('cruz_frac_manual')
            cf = cf if cf is not None else meta.get('cruz_frac', 0.20)
            af = meta.get('anca_frac_manual')
            af = af if af is not None else meta.get('anca_frac', 0.25)
            cf = _clamp(cf, 0.0, 0.5)
            af = _clamp(af, 0.0, 0.5)
            xc = xf - cf * L if fr else xf + cf * L
            xa = xr + af * L if fr else xr - af * L

            dist = abs(xa - xc)
            xmid = (xc + xa) / 2.0
            xlo = xmid - FRAC * dist
            xhi = xmid + FRAC * dist
            yfull = V[:, 1].min() - 1.0  # piso por debajo de todo -> sección completa
            vol = clipped_volume_liters(V, F, xlo, xhi, yfull, 24)

            rows.append({
                'dataset': lab,
                'individuo': name,
                'dist_cruz_anca_cm': round(dist, 1),
                'ancho_franja_cm': round(2 * FRAC * dist, 1),   # 2/8·dist = 1/4·dist
                'vol_central_L': round(vol, 1),
            })
            print(f"  {lab:8} {name:14} dist={dist:6.1f}cm  franja={2*FRAC*dist:5.1f}cm  vol={vol:6.1f}L")

    cols = ['dataset', 'individuo', 'dist_cruz_anca_cm', 'ancho_franja_cm', 'vol_central_L']
    with open('tabla_volumen_central.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)
    print(f"\n[done] {len(rows)} individuos -> tabla_volumen_central.csv")

    # Excel formateado
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = 'Vol central ±1-8 dist'
        hd = ['Dataset', 'Individuo', 'Dist cruz-anca (cm)', 'Ancho franja 2/8·dist (cm)', 'Vol central (L)']
        hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='D9D9D9'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.append(hd)
        for j in range(1, len(hd) + 1):
            c = ws.cell(1, j); c.fill = hf; c.font = hfont; c.border = bd
            c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        for r in rows:
            ws.append([r['dataset'], r['individuo'], r['dist_cruz_anca_cm'],
                       r['ancho_franja_cm'], r['vol_central_L']])
        for rr in range(2, len(rows) + 2):
            for cc in range(1, 6):
                cell = ws.cell(rr, cc); cell.border = bd
                if cc >= 3:
                    cell.number_format = '0.0'
        for j, w_ in enumerate([10, 16, 17, 22, 14], 1):
            ws.column_dimensions[get_column_letter(j)].width = w_
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:E{len(rows) + 1}"
        wb.save('tabla_volumen_central.xlsx')
        print('[done] tabla_volumen_central.xlsx')
    except Exception as e:
        print('[warn] xlsx omitido:', e)


if __name__ == '__main__':
    main()
