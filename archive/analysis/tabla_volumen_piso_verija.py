"""
Tabla del VOLUMEN cruz↔anca hasta el PISO DE LA VERIJA guardado por modelo.

Replica exactamente el "Vol. piso verija" del visor 3D (viewer3d.js):
  - Región a lo largo de X entre el plano de la CRUZ y el de la ANCA.
  - Piso horizontal a la altura de la verija GUARDADA:
      * verija_frac_manual (>0.25 default) ubica la sección de la verija (desde el fondo)
      * verija_raise_manual (>0.5 default) sube/baja el piso:
          yLow  = yVerBot - (yVerCap - yVerBot)     # 0%  (abajo, afuera del barril)
          yFloor = yLow + raise * (yVerCap - yLow)   # 0.5 = fondo de la panza, 1 = tope
  - Volumen = integral del área de cada sección recortada al piso, sobre X.

Sobre las mallas _v8 SIN cresta (las que muestra el visor). Salida: CSV + Excel.
Uso: python tabla_volumen_piso_verija.py
"""
import csv
import glob
import os
import json
import numpy as np
import trimesh
from tabla_volumen_corte import slice_section, section_info, clipped_volume_liters, V8

OCULTOS = {'000_392', '000_392A', '000_448A', '000_459'}  # 12junio ocultos


def _clamp(v, lo, hi):
    return min(max(float(v), lo), hi)


def _slice_retry(V, F, x0, step_dir, L):
    """Sección en X=x0; si vacía, empuja hacia el cuerpo y reintenta (como el visor)."""
    x = x0
    s = slice_section(V, F, x)
    k = 0
    while s is None and k < 4:
        x += step_dir * max(0.5, L * 0.01)
        s = slice_section(V, F, x)
        k += 1
    return s


def main():
    rows = []
    for lab, d in V8.items():
        for ind in sorted(glob.glob(d + '/*')):
            if not os.path.isdir(ind):
                continue
            name = os.path.basename(ind)
            if lab == '12junio' and name in OCULTOS:
                continue
            ply = glob.glob(ind + '/*_3d.ply')
            rj = glob.glob(ind + '/*_resumen.json')
            if not ply or not rj:
                continue
            meta = json.load(open(rj[0]))
            m = trimesh.load(ply[0], process=False)
            V = np.asarray(m.vertices, float)
            F = np.asarray(m.faces, int)
            if len(F) == 0:
                continue
            x = V[:, 0]
            xmin, xmax = x.min(), x.max()
            L = xmax - xmin
            fr = (meta.get('barril_dir') != 'left')
            xf = xmax if fr else xmin   # frente
            xr = xmin if fr else xmax   # fondo

            cf = meta.get('cruz_frac_manual')
            cf = cf if cf is not None else meta.get('cruz_frac', 0.20)
            af = meta.get('anca_frac_manual')
            af = af if af is not None else meta.get('anca_frac', 0.25)
            cf = _clamp(cf, 0.0, 0.5)
            af = _clamp(af, 0.0, 0.5)
            xc = xf - cf * L if fr else xf + cf * L
            xa = xr + af * L if fr else xr - af * L

            # Piso de verija GUARDADO: si no está, se descarta el animal.
            vf_saved = meta.get('verija_frac_manual')
            vr_saved = meta.get('verija_raise_manual')
            if vf_saved is None:
                print(f"  [skip] {lab}/{name}: sin piso de verija guardado")
                continue
            vf = _clamp(vf_saved, 0.0, 0.5)
            vr = _clamp(vr_saved if vr_saved is not None else 0.5, 0.0, 1.0)

            xver = xr + vf * L if fr else xr - vf * L
            secv = _slice_retry(V, F, xver, (1 if fr else -1), L)

            xlo, xhi = min(xc, xa), max(xc, xa)
            dist = xhi - xlo
            if secv is None:
                yfloor = V[:, 1].min() - 1.0  # sin verija -> sección completa
            else:
                iv = section_info(secv)
                yb, yc = iv['ymin'], iv['ymax']
                ylow = yb - (yc - yb)
                yfloor = ylow + vr * (yc - ylow)
            vol = clipped_volume_liters(V, F, xlo, xhi, yfloor, 36)

            rows.append({
                'individuo': name,
                'vol_piso_verija_L': round(vol, 1),
            })
            print(f"  {lab:8} {name:14}  vol={vol:6.1f}L")

    cols = ['individuo', 'vol_piso_verija_L']
    with open('tabla_volumen_piso_verija.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader(); w.writerows(rows)
    print(f"\n[done] {len(rows)} individuos -> tabla_volumen_piso_verija.csv")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook(); ws = wb.active; ws.title = 'Vol piso verija'
        hd = ['Individuo', 'Vol piso verija (L)']
        hf = PatternFill('solid', fgColor='C55A11'); hfont = Font(bold=True, color='FFFFFF')
        thin = Side(style='thin', color='D9D9D9'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.append(hd)
        for j in range(1, len(hd) + 1):
            c = ws.cell(1, j); c.fill = hf; c.font = hfont; c.border = bd
            c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
        for r in rows:
            ws.append([r[c] for c in cols])
        for rr in range(2, len(rows) + 2):
            for cc in range(1, len(cols) + 1):
                cell = ws.cell(rr, cc); cell.border = bd
                if cc == 2:
                    cell.number_format = '0.0'
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:B{len(rows) + 1}"
        wb.save('tabla_volumen_piso_verija.xlsx')
        print('[done] tabla_volumen_piso_verija.xlsx')
    except Exception as e:
        print('[warn] xlsx omitido:', e)


if __name__ == '__main__':
    main()
