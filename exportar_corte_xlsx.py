"""Convierte output_cruz_modelos/tabla_corte_barrido.csv -> Excel formateado.
Una fila por individuo; columnas de volumen de corte 40%..70% (cruz↔anca, sin cresta).
Uso: python exportar_corte_xlsx.py
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV = 'output_cruz_modelos/tabla_corte_barrido.csv'
OUT = 'tabla_corte_barrido_sincresta.xlsx'

rows = list(csv.DictReader(open(CSV)))
cols = list(rows[0].keys())  # respeta el orden del CSV

wb = Workbook(); ws = wb.active; ws.title = 'Corte 40-70% (sin cresta)'
hfill = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
thin = Side(style='thin', color='D9D9D9'); border = Border(left=thin, right=thin, top=thin, bottom=thin)

# encabezado legible
def nice(c):
    if c.startswith('vol_') and c.endswith('pct_L'):
        return c.replace('vol_', '').replace('pct_L', '%')
    return {'nombre': 'Individuo', 'dataset': 'Dataset', 'altura_calc_cm': 'Altura calc (cm)',
            'dist_cruz_anca_cm': 'Dist cruz-anca (cm)', 'diam_cruz_cm': 'Perím cruz (cm)',
            'diam_anca_cm': 'Perím anca (cm)'}.get(c, c)

ws.append([nice(c) for c in cols])
for j in range(1, len(cols) + 1):
    cell = ws.cell(1, j); cell.fill = hfill; cell.font = hfont; cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')

numcols = set(i + 1 for i, c in enumerate(cols)
              if c.startswith('vol_') or c in ('altura_calc_cm', 'dist_cruz_anca_cm', 'diam_cruz_cm', 'diam_anca_cm'))
for r in rows:
    ws.append([r[c] for c in cols])
nrows = len(rows) + 1
for rr in range(2, nrows + 1):
    for cc in range(1, len(cols) + 1):
        cell = ws.cell(rr, cc); cell.border = border
        if cc in numcols:
            try:
                cell.value = float(cell.value); cell.number_format = '0.0'
            except (TypeError, ValueError):
                pass

# anchos: primeras 2 anchas, resto angostas
for j, c in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(j)].width = 16 if j <= 2 else (13 if not c.startswith('vol_') else 7)
ws.freeze_panes = 'C2'  # fija encabezado + individuo/dataset
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{nrows}"

wb.save(OUT)
print(f"guardado: {OUT}  ({len(rows)} individuos, {sum(c.startswith('vol_') for c in cols)} columnas 40-70%)")
